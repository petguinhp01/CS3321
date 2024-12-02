"""
File: data_manager.py
Description: Centralized management for unified data file operations.
Authors: Nhan Nguyen (Final Revised)
Date: 12/01/2024
"""

from openpyxl import Workbook, load_workbook
import os

DATA_FILE = "unified_data.xlsx"

def initialize_data_file():
    """
    Initializes the unified Excel file with required sheets if it doesn't exist.
    """
    try:
        if not os.path.exists(DATA_FILE):
            wb = Workbook()

            # Define the required sheets and their headers
            sheets = {
                "User": ["Username", "Password", "Role"],
                "Staff": ["Name", "Role", "Email"],
                "Services": ["Service Name", "Category", "Duration", "Price"],
                "Bookings": ["Booking ID", "Customer Name", "Service", "Date", "Time", "Staff Name", "Status"],
                "Staff Availability": ["Staff Name", "Day", "Time Slot"],
            }

            for idx, (sheet_name, headers) in enumerate(sheets.items()):
                ws = wb.active if idx == 0 else wb.create_sheet(sheet_name)
                ws.title = sheet_name
                ws.append(headers)

            wb.save(DATA_FILE)
            print(f"Database file created at {DATA_FILE}.")
    except Exception as e:
        print(f"Error initializing data file: {e}")


def get_sheet(sheet_name):
    """
    Retrieves the specified sheet from the unified Excel file.

    Parameters:
        sheet_name (str): The name of the sheet to retrieve.

    Returns:
        Workbook, Worksheet: The workbook and the specified sheet object.
    """
    try:
        wb = load_workbook(DATA_FILE)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' does not exist in {DATA_FILE}.")
        return wb, wb[sheet_name]
    except Exception as e:
        print(f"Error accessing sheet '{sheet_name}': {e}")
        return None, None


def append_row(sheet_name, row_data):
    """
    Appends a row to the specified sheet.

    Parameters:
        sheet_name (str): The name of the sheet to update.
        row_data (list): A list of values to add as a new row.
    """
    try:
        wb, ws = get_sheet(sheet_name)
        if ws:
            ws.append(row_data)
            wb.save(DATA_FILE)
    except Exception as e:
        print(f"Error appending row to '{sheet_name}': {e}")


def read_all_rows(sheet_name):
    """
    Reads all rows from the specified sheet.

    Parameters:
        sheet_name (str): The name of the sheet to read from.

    Returns:
        list: A list of tuples containing all rows in the sheet.
    """
    try:
        _, ws = get_sheet(sheet_name)
        if ws:
            return [row for row in ws.iter_rows(min_row=2, values_only=True)]  # Skip header row
    except Exception as e:
        print(f"Error reading rows from '{sheet_name}': {e}")
    return []


def delete_row(sheet_name, column_index, value):
    """
    Deletes rows in the specified sheet where the column matches the value.

    Parameters:
        sheet_name (str): The name of the sheet to update.
        column_index (int): The column index (0-based) to match.
        value (str): The value to search for in the specified column.
    """
    try:
        wb, ws = get_sheet(sheet_name)
        if ws:
            rows_to_delete = [row[0].row for row in ws.iter_rows(min_row=2) if row[column_index].value == value]
            for row_idx in reversed(rows_to_delete):  # Delete rows in reverse order
                ws.delete_rows(row_idx)
            wb.save(DATA_FILE)
    except Exception as e:
        print(f"Error deleting row from '{sheet_name}': {e}")


def update_row(sheet_name, column_index, value, new_data):
    """
    Updates a row in the specified sheet where the column matches the value.

    Parameters:
        sheet_name (str): The name of the sheet to update.
        column_index (int): The column index (0-based) to match.
        value (str): The value to search for in the specified column.
        new_data (list): A list of new values to replace the row data.
    """
    try:
        wb, ws = get_sheet(sheet_name)
        if ws:
            for row in ws.iter_rows(min_row=2):
                if row[column_index].value == value:
                    for col_idx, new_value in enumerate(new_data, start=1):
                        ws.cell(row=row[0].row, column=col_idx).value = new_value
                    break
            wb.save(DATA_FILE)
    except Exception as e:
        print(f"Error updating row in '{sheet_name}': {e}")
