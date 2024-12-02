"""
File: utils.py
Description: Utility functions for shared operations like saving users and initializing files.
Authors: Nhan Nguyen (Revised)
Date: 12/01/2024
"""

import os
from openpyxl import load_workbook
from data_manager import initialize_data_file, DATA_FILE

def save_user(username, password, role):
    """
    Saves a user account to the 'User' sheet in the unified Excel file.

    Parameters:
        username (str): Username of the account.
        password (str): Password of the account.
        role (str): Role of the user (e.g., Admin, Staff, Customer).
    """
    try:
        # Ensure the data file is initialized
        initialize_data_file()

        # Load the workbook and check for the "User" sheet
        wb = load_workbook(DATA_FILE)
        if "User" not in wb.sheetnames:
            ws = wb.create_sheet("User")
            ws.append(["Username", "Password", "Role"])  # Add headers
        else:
            ws = wb["User"]

        # Append the user details
        ws.append([username, password, role])
        wb.save(DATA_FILE)
        print(f"User {username} added successfully.")
    except Exception as e:
        print(f"Error saving user: {e}")


def save_staff(name, role, email):
    """
    Saves a staff member to the 'Staff' sheet in the unified Excel file.

    Parameters:
        name (str): Name of the staff member.
        role (str): Role of the staff member (e.g., Manager, Staff).
        email (str): Email of the staff member.
    """
    try:
        # Ensure the data file is initialized
        initialize_data_file()

        # Load the workbook and check for the "Staff" sheet
        wb = load_workbook(DATA_FILE)
        if "Staff" not in wb.sheetnames:
            ws = wb.create_sheet("Staff")
            ws.append(["Name", "Role", "Email"])  # Add headers
        else:
            ws = wb["Staff"]

        # Append the staff details
        ws.append([name, role, email])
        wb.save(DATA_FILE)
        print(f"Staff member {name} added successfully.")
    except Exception as e:
        print(f"Error saving staff: {e}")


def save_service(service_name, category, duration, price):
    """
    Saves a service to the 'Services' sheet in the unified Excel file.

    Parameters:
        service_name (str): Name of the service.
        category (str): Category of the service.
        duration (str): Duration of the service (e.g., 1h 30m).
        price (float): Price of the service.
    """
    try:
        # Ensure the data file is initialized
        initialize_data_file()

        # Load the workbook and check for the "Services" sheet
        wb = load_workbook(DATA_FILE)
        if "Services" not in wb.sheetnames:
            ws = wb.create_sheet("Services")
            ws.append(["Service Name", "Category", "Duration", "Price"])  # Add headers
        else:
            ws = wb["Services"]

        # Append the service details
        ws.append([service_name, category, duration, price])
        wb.save(DATA_FILE)
        print(f"Service {service_name} added successfully.")
    except Exception as e:
        print(f"Error saving service: {e}")


def save_availability(staff_name, day, time_slot):
    """
    Saves an availability slot for a staff member in the 'Staff Availability' sheet.

    Parameters:
        staff_name (str): Name of the staff member.
        day (str): Day of the week (e.g., Monday).
        time_slot (str): Time slot (e.g., 10:00-14:00).
    """
    try:
        # Ensure the data file is initialized
        initialize_data_file()

        # Load the workbook and check for the "Staff Availability" sheet
        wb = load_workbook(DATA_FILE)
        if "Staff Availability" not in wb.sheetnames:
            ws = wb.create_sheet("Staff Availability")
            ws.append(["Staff Name", "Day", "Time Slot"])  # Add headers
        else:
            ws = wb["Staff Availability"]

        # Append the availability slot
        ws.append([staff_name, day, time_slot])
        wb.save(DATA_FILE)
        print(f"Availability added for {staff_name} on {day} during {time_slot}.")
    except Exception as e:
        print(f"Error saving availability: {e}")
